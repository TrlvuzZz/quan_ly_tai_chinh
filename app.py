import os
import io
import subprocess
import sys

# ── auto-install dependencies on first run ──────────────────────────────────
def ensure_deps():
    req = os.path.join(os.path.dirname(__file__), 'requirements.txt')
    try:
        import flask
        import openpyxl
    except ImportError:
        print("📦 Đang cài đặt thư viện cần thiết...")
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', '-r', req])
        print("✅ Cài đặt xong!\n")

ensure_deps()
# ────────────────────────────────────────────────────────────────────────────

from flask import Flask, jsonify, request, render_template, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import database as db

app = Flask(__name__)

# ── Init DB on startup ───────────────────────────────────────────────────────
db.init_db()


# ── Pages ────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')


# ── API: Transactions ────────────────────────────────────────────────────────
@app.route('/api/transactions', methods=['GET'])
def api_get_all():
    return jsonify(db.get_all())


@app.route('/api/transactions', methods=['POST'])
def api_create():
    d = request.json
    missing = [f for f in ('date', 'category', 'type', 'amount') if not d.get(f)]
    if missing:
        return jsonify({'error': f'Thiếu trường: {", ".join(missing)}'}), 400
    if d['type'] not in ('Thu', 'Chi'):
        return jsonify({'error': 'type phải là Thu hoặc Chi'}), 400
    try:
        amount = float(d['amount'])
        if amount <= 0:
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({'error': 'Số tiền không hợp lệ'}), 400
    row = db.create(d['date'], d['category'], d['type'], amount, d.get('note', ''))
    return jsonify(row), 201


@app.route('/api/transactions/<int:tx_id>', methods=['PUT'])
def api_update(tx_id):
    if not db.get_one(tx_id):
        return jsonify({'error': 'Không tìm thấy giao dịch'}), 404
    d = request.json
    try:
        amount = float(d['amount'])
        if amount <= 0:
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({'error': 'Số tiền không hợp lệ'}), 400
    row = db.update(tx_id, d['date'], d['category'], d['type'], amount, d.get('note', ''))
    return jsonify(row)


@app.route('/api/transactions/<int:tx_id>', methods=['DELETE'])
def api_delete(tx_id):
    if not db.get_one(tx_id):
        return jsonify({'error': 'Không tìm thấy giao dịch'}), 404
    db.delete(tx_id)
    return jsonify({'ok': True})


@app.route('/api/stats', methods=['GET'])
def api_stats():
    return jsonify(db.get_stats())


# ── Export Excel ─────────────────────────────────────────────────────────────
@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    rows = db.get_all()
    wb = openpyxl.Workbook()

    # ── Sheet 1: Danh sách giao dịch ────────────────────────────────────────
    ws = wb.active
    ws.title = 'Giao Dịch'

    header_fill  = PatternFill('solid', fgColor='1E2535')
    green_fill   = PatternFill('solid', fgColor='D1FAE5')
    red_fill     = PatternFill('solid', fgColor='FEE2E2')
    header_font  = Font(bold=True, color='E2E8F0', name='Calibri', size=11)
    normal_font  = Font(name='Calibri', size=10)
    center       = Alignment(horizontal='center', vertical='center')
    thin         = Side(style='thin', color='CBD5E0')
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['ID', 'Ngày', 'Danh Mục', 'Loại', 'Số Tiền (VNĐ)', 'Ghi Chú']
    col_widths = [8, 14, 22, 10, 20, 30]

    ws.row_dimensions[1].height = 28
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(rows, 2):
        vals = [row['id'], row['date'], row['category'], row['type'],
                row['amount'], row['note']]
        fill = green_fill if row['type'] == 'Thu' else red_fill
        ws.row_dimensions[ri].height = 20
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center
            if ci in (1, 4):
                cell.fill = fill
            if ci == 5:
                cell.number_format = '#,##0'

    # ── Sheet 2: Báo cáo theo tháng ─────────────────────────────────────────
    ws2 = wb.create_sheet('Báo Cáo Tháng')
    months = sorted({r['date'][:7] for r in rows}, reverse=True)

    ws2.row_dimensions[1].height = 28
    h2 = ['Tháng', 'Tổng Thu', 'Tổng Chi', 'Số Dư', 'Số GD']
    for ci, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    for w, ci in zip([14, 20, 20, 20, 10], range(1, 6)):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    for ri, m in enumerate(months, 2):
        m_rows = [r for r in rows if r['date'].startswith(m)]
        thu  = sum(r['amount'] for r in m_rows if r['type'] == 'Thu')
        chi  = sum(r['amount'] for r in m_rows if r['type'] == 'Chi')
        net  = thu - chi
        vals = [m, thu, chi, net, len(m_rows)]
        ws2.row_dimensions[ri].height = 20
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center
            if ci in (2, 3, 4):
                cell.number_format = '#,##0'
            if ci == 4:
                cell.font = Font(name='Calibri', size=10, bold=True,
                                 color='059669' if net >= 0 else 'DC2626')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import date
    fname = f'GiaoDich_{date.today().isoformat()}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


# ── Run ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("=" * 55)
    print("  💰  Quản Lý Tài Chính Cá Nhân")
    print("  🌐  http://127.0.0.1:5000")
    print("  📁  Database:", db.DB_PATH)
    print("=" * 55)
    app.run(debug=True, port=5000)
